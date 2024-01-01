from random import randint
from openpyxl import load_workbook

def fazer_jogos():
    lista = list()
    jogos = list()

    workbook = load_workbook(filename='resultados_mega_sena.xlsx')
    sheet = workbook.active

    sheet['A1'] = '-'*60
    sheet['A2'] = '-'*25 + ' MEGA SENA ' + '-'*24
    sheet['A3'] = '-'*60

    quant = int(input('Quantos jogos foram realizados: '))
    tot = 1

    while tot <= quant:
        cont = 0
        while True:
            num = randint(1, 100)  # 1 até 60
            if num not in lista:
                lista.append(num)
                cont += 1
            if cont >= 20:
                break
        lista.sort()
        jogos.append(lista[:])
        lista.clear()
        tot += 1

    sheet['A4'] = '-=' * 3 + f' SORTEANDO {quant} JOGOS ' + '-=' * 3
    for i, l in enumerate(jogos):
        l.sort()  # Ordenando os números de cada jogo
        cell = sheet.cell(row=i+5, column=1)
        cell.value = f' Jogo {i+1}: {l}'

    sheet['A' + str(len(jogos) + 5)] = '-=' * 5 + ' < BOA SORTE! > ' + '-=' * 5

    workbook.save(filename='resultados_mega_sena.xlsx')

def verificar_numeros():
    workbook = load_workbook(filename='resultados_mega_sena.xlsx')
    sheet = workbook.active

    numeros_reais = input('Digite os números sorteados separados por vírgula: ')
    numeros_reais = list(map(int, numeros_reais.split(',')))

    numeros_reais.sort()
    encontrado = False
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        jogo = list(map(int, row[0].split(':')[1].strip().strip('[]').split(', ')))
        jogo.sort()
        if numeros_reais == jogo:
            print(f'Os números reais foram encontrados no Jogo {sheet.cell(row=row[0].row, column=1).row - 4}!')
            encontrado = True
            break

    if not encontrado:
        print('Os números reais não correspondem a nenhum jogo gerado.')

def main():
    while True:
        print("\nMENU:")
        print("1. Fazer Jogos")
        print("2. Verificar Números")
        print("3. Sair")
        escolha = input("Escolha uma opção: ")

        if escolha == '1':
            fazer_jogos()
        elif escolha == '2':
            verificar_numeros()
        elif escolha == '3':
            break
        else:
            print("Opção inválida. Por favor, escolha uma opção válida.")

if __name__ == "__main__":
    main()
